# export_account_data_fun.py

# 导出科目余额表、序时账 xlsx 文件

from tkinter import filedialog
import openpyxl
from openpyxl.utils import get_column_letter
from xlsx_tools_openxl import cell_format_fun

def export_account_data(df_balance, df_chronological):

    wb = openpyxl.Workbook()

    # 查找表
    ws_find = wb.active
    ws_find.title = '查找'

    cell_A1 = ws_find.cell(row=1, column=1, value='科目代码')
    cell_format_fun.cell_format(cell_A1, 2)

    cell_B1 = ws_find.cell(row=1, column=2, value='')
    cell_B1.number_format = '@'
    cell_format_fun.cell_format(cell_B1, 2)

    ws_find.cell(row=3, column=1, value='科目余额表')               # A3

    cell_A4 = ws_find.cell(row=4, column=1, value='科目编码')       # A4
    cell_A4.number_format = '@'

    ws_find.cell(row=4, column=2, value='科目名称')                 # B4
    ws_find.cell(row=4, column=3, value='期初借方')                 # C4
    ws_find.cell(row=4, column=4, value='期初贷方')                 # D4
    ws_find.cell(row=4, column=5, value='本期借方')                 # E4
    ws_find.cell(row=4, column=6, value='本期贷方')                 # F4
    ws_find.cell(row=4, column=7, value='期末借方')                 # G4
    ws_find.cell(row=4, column=8, value='期末贷方')                 # H4

    for row in ws_find.iter_rows(min_row=4, max_row=5, min_col=1, max_col=8):
        for cell in row:
            cell_format_fun.cell_format(cell, 2)

    ws_find.cell(row=7, column=1, value='序时账')                  # A7

    ws_find.cell(row=7, column=1, value='涉及科目')                 # A7
    ws_find.cell(row=7, column=2, value='日期')                    # B7
    ws_find.cell(row=7, column=3, value='凭证字号')                 # C7
    ws_find.cell(row=7, column=4, value='科目编码')                 # D7
    ws_find.cell(row=7, column=5, value='科目名称')                 # E7
    ws_find.cell(row=7, column=6, value='摘要')                    # F7
    ws_find.cell(row=7, column=7, value='对方科目')                 # G7
    ws_find.cell(row=7, column=8, value='借方金额')                 # H7
    ws_find.cell(row=7, column=9, value='贷方金额')                 # I7
    ws_find.cell(row=7, column=10, value='现金流量项目')             # J7

    for row in ws_find.iter_rows(min_row=7, max_row=10007, min_col=1, max_col=8):
        for cell in row:
            cell_format_fun.cell_format(cell, 2)

    ws_find.auto_filter.ref = "A7:I7"

    # 设置日期和数值格式，遍历所有列
    for col in ws_find.iter_cols(min_row=8, max_row=10007, min_col=2, max_col=2):
        for cell in col:
            cell.number_format = 'yyyy-mm-dd'

    for col in ws_find.iter_cols(min_row=8, max_row=10007, min_col=7, max_col=8):
        for cell in col:
            cell.number_format = '#,##0.00'             # 数字格式设置为千分位和保留两位小数

    # 调整列宽（按实际列顺序）
    find_widths = {
                    'A': 15,
                    'B': 15,
                    'C': 15,
                    'D': 25,
                    'E': 50,
                    'F': 50,
                    'G': 15,
                    'H': 15,
                    'I': 15,
                    'J': 15,
    }

    # 设置列宽
    for i, (col_name, width) in enumerate(find_widths.items(), start=1):
        col_letter = get_column_letter(i)
        ws_find.column_dimensions[col_letter].width = width


    # 科目余额表 
    ws_balance = wb.create_sheet(title='科目余额表')

    for j, col in enumerate(df_balance.columns):
        ws_balance.cell(row=1, column=1+j, value=col)

    for i, row in enumerate(df_balance.values):
        for j, value in enumerate(row):
            ws_balance.cell(row=2+i, column=1+j, value=value)

    for row in ws_balance.iter_rows(min_row=1, max_row=ws_balance.max_row, min_col=1, max_col=ws_balance.max_column):
        for cell in row:
            cell_format_fun.cell_format(cell, 2)

    # 添加筛选功能，筛选行设置在第一行
    ws_balance.auto_filter.ref = ws_balance.dimensions

    # 设置数值格式，遍历所有列
    for col in ws_balance.iter_cols(min_row=2, max_row=ws_balance.max_row, min_col=2, max_col=ws_balance.max_column):
        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell_format_fun.cell_format(cell, 4)             # 数字格式设置为千分位和保留两位小数

    # 调整列宽（按实际列顺序）
    balance_widths = {
        '科目编号': 15,
        '科目名称': 50,
        '期初借方': 15,
        '期初贷方': 15,
        '本期借方': 15,
        '本期贷方': 15,
        '期末借方': 15,
        '期末贷方': 15
    }

    # 设置列宽
    for i, (col_name, width) in enumerate(balance_widths.items(), start=1):
        col_letter = get_column_letter(i)
        ws_balance.column_dimensions[col_letter].width = width


    # 序时账
    ws_chronological = wb.create_sheet(title='序时账')

    for j, col in enumerate(df_chronological.columns):
        ws_chronological.cell(row=1, column=1+j, value=col)

    for i, row in enumerate(df_chronological.values):
        for j, value in enumerate(row):
            ws_chronological.cell(row=2+i, column=1+j, value=value)

    for row in ws_chronological.iter_rows(min_row=1, max_row=ws_chronological.max_row, min_col=1, max_col=ws_chronological.max_column):
        for cell in row:
            cell_format_fun.cell_format(cell, 2)

    # 添加筛选功能，筛选行设置在第一行
    ws_chronological.auto_filter.ref = ws_chronological.dimensions

    # 设置日期和数值格式，遍历所有列
    for col in ws_chronological.iter_cols(min_row=2, max_row=ws_chronological.max_row, min_col=2, max_col=2):
        for cell in col:
            cell_format_fun.cell_format(cell, 5)

    for col in ws_chronological.iter_cols(min_row=2, max_row=ws_chronological.max_row, min_col=7, max_col=ws_chronological.max_column):
        for cell in col:
            cell_format_fun.cell_format(cell, 4)             # 数字格式设置为千分位和保留两位小数

    # 调整列宽（按实际列顺序）
    chronological_widths = {
                            '涉及科目': 10,
                            '日期': 15,
                            '凭证字号': 15,
                            '科目编码': 30,
                            '科目名称': 50,
                            '摘要': 50,
                            '对方科目': 15,
                            '借方金额': 15,
                            '贷方金额': 15,
                            '现金流量项目': 20
    }

    # 设置列宽
    for i, (col_name, width) in enumerate(chronological_widths.items(), start=1):
        col_letter = get_column_letter(i)
        ws_chronological.column_dimensions[col_letter].width = width


    # 保存文件
    path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel Files', '*.xlsx')])

    if path:
    
        wb.save(path)

        info = 'A5单元格公式：=XLOOKUP(B1,科目余额表!A:A,科目余额表!A:H)\n'
        info += 'A8单元格公式：=FILTER(序时账!A:H,ISNUMBER(SEARCH(B1,序时账!A:A)))\n'
        info += f'Path: {path}\nExport successful!\n'

        return [True, info, path]
    
    else:

        info = 'Export failed!\n'

        return [False, info]