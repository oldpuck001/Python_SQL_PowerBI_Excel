# gui_tk_fill_info_sheet.py

import os
import json
import shutil
import openpyxl
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from gui_tk import gui_tk_area_text
from dataframe_tools_pd import read_xlsx_xls_csv_txt_fun

class gui_tk_fill_info_sheet_class:

    info_dict = {'企业名称': '', '成立日期': '', '核准日期': '', '统一社会信用代码': '', '注册资本': '', '企业类型': '',
                 '法定代表人': '', '注册地址': '', '国标行业': '', '登记机关': '', '经营范围': '',
                 '股东1名称': '', '股东1性质': '', '股东1注册资本': '', '股东1实收资本': '',
                 '股东2名称': '', '股东2性质': '', '股东2注册资本': '', '股东2实收资本': '',
                 '股东3名称': '', '股东3性质': '', '股东3注册资本': '', '股东3实收资本': '',
                 '股东4名称': '', '股东4性质': '', '股东4注册资本': '', '股东4实收资本': '',
                 '股东5名称': '', '股东5性质': '', '股东5注册资本': '', '股东5实收资本': '',
                 '股东6名称': '', '股东6性质': '', '股东6注册资本': '', '股东6实收资本': '',
                 '股东7名称': '', '股东7性质': '', '股东7注册资本': '', '股东7实收资本': ''
                 }

    def gui_tk_fill_info_sheet_frame(self, root, control_frame_config, text_area):

        frame_result = tk.Frame(root)
        frame_result.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=(10, 5))

        # 按钮
        frame_result.frame_button = tk.Frame(frame_result)
        frame_result.frame_button.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Button(frame_result.frame_button, text=control_frame_config['widget_text'][0],
                  command=lambda: self.import_file(text_area),
                  width=25).pack(side=tk.LEFT, padx=(50, 5), pady=(10, 5))
        
        tk.Button(frame_result.frame_button, text=control_frame_config['widget_text'][1],
                  command=lambda: self.export_file(text_area),
                  width=25).pack(side=tk.LEFT, padx=5, pady=(10, 5))

        return frame_result
    

    def import_file(self, text_area):

        fill_text = ''
        path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx'),
                                                     ('Excel Files', '*.xls')])

        # 读入数据
        result_info = read_xlsx_xls_csv_txt_fun.read_xlsx_xls_csv_txt(file_path=path, sheet_name='基本信息')

        if result_info[0]:

            basic_info_df = result_info[2]
            shareholder_name = ['', '', '', '', '', '', '']
            shareholder_type = ['', '', '', '', '', '', '']
            shareholder_registered = ['', '', '', '', '', '', '']
            shareholder_paid = ['', '', '', '', '', '', '']

            # 提取所需信息
            for i in range(11):
                for j in range(1, len(basic_info_df)):
                    keywords = basic_info_df.iloc[j, i]
                    if keywords == '企业名称' and j < 18:
                        enterprise_name = basic_info_df.iloc[j, i+1]               # 企业名称
                    elif keywords == '成立日期' and j < 18:
                        date_of_establishment = basic_info_df.iloc[j, i+1]         # 成立日期
                    elif keywords == '核准日期' and j < 18:
                        approval_date = basic_info_df.iloc[j, i+1]                 # 核准日期
                    elif keywords == '统一社会信用代码' and j < 18:
                        unified_social_credit_code = basic_info_df.iloc[j, i+1]    # 统一社会信用代码
                    elif keywords == '注册资本' and j < 18:
                        registered_capital = basic_info_df.iloc[j, i+1]            # 注册资本
                    elif keywords == '企业类型' and j < 18:
                        enterprise_type = basic_info_df.iloc[j, i+1]               # 企业类型
                    elif keywords == '法定代表人' and j < 18:
                        legal_representative = basic_info_df.iloc[j, i+1]          # 法定代表人
                    elif keywords == '注册地址' and j < 18 and j < 18:
                        registered_address = basic_info_df.iloc[j, i+1]            # 注册地址
                    elif keywords == '国标行业' and j < 18:
                        national_standard_industry = basic_info_df.iloc[j, i+1]    # 国标行业
                    elif keywords == '登记机关' and j < 18:
                        registration_authority = basic_info_df.iloc[j, i+1]        # 登记机关
                    elif keywords == '经营范围' and j < 18:
                        business_scope_original = basic_info_df.iloc[j, i+1]       # 经营范围
                        business_scope = business_scope_original.replace('\n', '').replace('\r', '')
                    elif keywords == '股东名称' and j < 23:
                        for n in range(7):
                            value = basic_info_df.iloc[j+n+1, i]
                            if pd.notna(value):
                                shareholder_name[n] = value
                            else:
                                break
                    elif keywords == '股东类型' and j < 23:
                        for n in range(7):
                            value = basic_info_df.iloc[j+n+1, i]
                            if pd.notna(value):
                                shareholder_type[n] = value
                            else:
                                break
                    elif keywords == '认缴出资额' and j < 23:
                        for n in range(7):
                            value = basic_info_df.iloc[j+n+1, i]
                            if pd.notna(value):
                                shareholder_registered[n] = float(value.replace('（万元）', ''))
                            else:
                                break
                    elif keywords == '实缴出资额' and j < 23:
                        for n in range(7):
                            value = basic_info_df.iloc[j+n+1, i]
                            if pd.notna(value):
                                shareholder_paid[n] = float(value.replace('（万元）', ''))
                            else:
                                break

            # 创建一个字典来保存这些信息
            self.info_dict = {
                                '企业名称': enterprise_name,
                                '成立日期': date_of_establishment,
                                '核准日期': approval_date,
                                '统一社会信用代码': unified_social_credit_code,
                                '注册资本': registered_capital,
                                '企业类型': enterprise_type,
                                '法定代表人': legal_representative,
                                '注册地址': registered_address,
                                '国标行业': national_standard_industry,
                                '登记机关': registration_authority,
                                '经营范围': business_scope,
                                '股东1名称': shareholder_name[0],
                                '股东1性质': shareholder_type[0],
                                '股东1注册资本': shareholder_registered[0],
                                '股东1实收资本': shareholder_paid[0],
                                '股东2名称': shareholder_name[1],
                                '股东2性质': shareholder_type[1],
                                '股东2注册资本': shareholder_registered[1],
                                '股东2实收资本': shareholder_paid[1],
                                '股东3名称': shareholder_name[2],
                                '股东3性质': shareholder_type[2],
                                '股东3注册资本': shareholder_registered[2],
                                '股东3实收资本': shareholder_paid[2],
                                '股东4名称': shareholder_name[3],
                                '股东4性质': shareholder_type[3],
                                '股东4注册资本': shareholder_registered[3],
                                '股东4实收资本': shareholder_paid[3],
                                '股东5名称': shareholder_name[4],
                                '股东5性质': shareholder_type[4],
                                '股东5注册资本': shareholder_registered[4],
                                '股东5实收资本': shareholder_paid[4],
                                '股东6名称': shareholder_name[5],
                                '股东6性质': shareholder_type[5],
                                '股东6注册资本': shareholder_registered[5],
                                '股东6实收资本': shareholder_paid[5],
                                '股东7名称': shareholder_name[6],
                                '股东7性质': shareholder_type[6],
                                '股东7注册资本': shareholder_registered[6],
                                '股东7实收资本': shareholder_paid[6],
                            }

            fill_text += 'Data read successful!\n'

        else:

            fill_text += 'Data read failed!\n'

        gui_tk_area_text.text_area_fill(text_area, fill_text)


    def export_file(self, text_area):

        fill_text = ''

        current_script_dir = os.path.dirname(os.path.abspath(__file__))
        source_file_path = os.path.join(current_script_dir, '..', 'xlsx_file', 'basic_info_sheet.xlsx')
        target_file_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel Files', '*.xlsx')])

        if target_file_path:

            shutil.copy(source_file_path, target_file_path)

            wb = openpyxl.load_workbook(target_file_path , data_only=False)
            ws = wb['基本情况表']

            # D3 企业名称
            ws.cell(row=3, column=4, value=self.info_dict['企业名称'])
            # G3 法定代表人
            ws.cell(row=3, column=7, value=self.info_dict['法定代表人'])
            # D4 注册地址
            ws.cell(row=4, column=4, value=self.info_dict['注册地址'])
            # I4 国标行业
            ws.cell(row=4, column=9, value=self.info_dict['国标行业'])
            # D5 企业类型
            ws.cell(row=5, column=4, value=self.info_dict['企业类型'])
            # G5 经营范围
            ws.cell(row=5, column=7, value=self.info_dict['经营范围'])
            # D8 登记机关
            ws.cell(row=8, column=4, value=self.info_dict['登记机关'])
            # I8 成立日期
            ws.cell(row=8, column=9, value=self.info_dict['成立日期'])
            # C15 股东1 名称
            ws.cell(row=15, column=3, value=self.info_dict['股东1名称'])
            # D15 股东1 性质
            ws.cell(row=15, column=4, value=self.info_dict['股东1性质'])
            # E15 股东1 注册资本
            ws.cell(row=15, column=5, value=self.info_dict['股东1注册资本'])
            # G15 股东1 实收资本
            ws.cell(row=15, column=7, value=self.info_dict['股东1实收资本'])
            # C16 股东2 名称
            ws.cell(row=16, column=3, value=self.info_dict['股东2名称'])
            # D16 股东2 性质
            ws.cell(row=16, column=4, value=self.info_dict['股东2性质'])
            # E16 股东2 注册资本
            ws.cell(row=16, column=5, value=self.info_dict['股东2注册资本'])
            # G16 股东2 实收资本
            ws.cell(row=16, column=7, value=self.info_dict['股东2实收资本'])
            # C17 股东3 名称
            ws.cell(row=17, column=3, value=self.info_dict['股东3名称'])
            # D17 股东3 性质
            ws.cell(row=17, column=4, value=self.info_dict['股东3性质'])
            # E17 股东3 注册资本
            ws.cell(row=17, column=5, value=self.info_dict['股东3注册资本'])
            # G17 股东3 实收资本
            ws.cell(row=17, column=7, value=self.info_dict['股东3实收资本'])
            # C18 股东4 名称
            ws.cell(row=18, column=3, value=self.info_dict['股东4名称'])
            # D18 股东4 性质
            ws.cell(row=18, column=4, value=self.info_dict['股东4性质'])
            # E18 股东4 注册资本
            ws.cell(row=18, column=5, value=self.info_dict['股东4注册资本'])
            # G18 股东4 实收资本
            ws.cell(row=18, column=7, value=self.info_dict['股东4实收资本'])
            # C19 股东5 名称
            ws.cell(row=19, column=3, value=self.info_dict['股东5名称'])
            # D19 股东5 性质
            ws.cell(row=19, column=4, value=self.info_dict['股东5性质'])
            # E19 股东5 注册资本
            ws.cell(row=19, column=5, value=self.info_dict['股东5注册资本'])
            # G19 股东5 实收资本
            ws.cell(row=19, column=7, value=self.info_dict['股东5实收资本'])
            # C20 股东6 名称
            ws.cell(row=20, column=3, value=self.info_dict['股东6名称'])
            # D20 股东6 性质
            ws.cell(row=20, column=4, value=self.info_dict['股东6性质'])
            # E20 股东6 注册资本
            ws.cell(row=20, column=5, value=self.info_dict['股东6注册资本'])
            # G20 股东6 实收资本
            ws.cell(row=20, column=7, value=self.info_dict['股东6实收资本'])
            # C21 股东7 名称
            ws.cell(row=21, column=3, value=self.info_dict['股东7名称'])
            # D21 股东7 性质
            ws.cell(row=21, column=4, value=self.info_dict['股东7性质'])
            # E21 股东7 注册资本
            ws.cell(row=21, column=5, value=self.info_dict['股东7注册资本'])
            # G21 股东7 实收资本
            ws.cell(row=21, column=7, value=self.info_dict['股东7实收资本'])

            wb.save(target_file_path)

            fill_text += f'Path: {target_file_path}\nCreate successful!\n'

        else:

            fill_text += f'Path: {target_file_path}\nCreate failed!\n'

        gui_tk_area_text.text_area_fill(text_area, fill_text)