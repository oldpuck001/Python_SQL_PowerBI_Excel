# account_balance_chronological.py

# 数据规范科目余额表、序时账

import os
import tempfile
import sys
import subprocess
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter.scrolledtext import ScrolledText
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

class App:

    file_path_result = ''
    df_balance = pd.DataFrame(columns=['科目编码', '科目名称', '期初借方', '期初贷方', '本期借方', '本期贷方', '期末借方', '期末贷方'])
    df_chronological = pd.DataFrame(columns=['涉及科目', '日期', '凭证字号', '科目编码', '科目名称', '摘要', '借方金额', '贷方金额'])

    def __init__(self):

        self.root = tk.Tk()                                             # 创建tk实例

        self.root.title('数据规范科目余额表、序时账')                       # 设置窗口标题

        self.root.geometry('555x240+50+50')                             # 设置窗口的大小和位置

        self.root.resizable(False, False)                               # 设置窗口是否可以调整大小

        if sys.platform == 'darwin':
            self.root.after(200, self.bring_to_front)                   # macOS workaround: mainloop開始後再將視窗浮前

        # 按钮
        self.frame_button = tk.Frame(self.root)
        self.frame_button.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=(10, 5))
        tk.Button(self.frame_button, text='导入科目余额表',
                  command=self.account_balance,
                  width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(self.frame_button, text='导入序时账',
                  command=self.account_chronological,
                  width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(self.frame_button, text='输出数据文件',
                  command=lambda: self.export_file(self.text_area),
                  width=15).pack(side=tk.LEFT, padx=5)

        # 操作记录区
        self.frame_text_area = tk.Frame(self.root)
        self.frame_text_area.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
        tk.Label(self.frame_text_area, text='Operation Log', anchor='w').pack(side=tk.TOP, fill=tk.X, padx=5, pady=(0, 5))
        self.text_area = ScrolledText(self.frame_text_area, height=10)
        self.text_area.pack(side=tk.TOP, expand=True, fill=tk.X)
        self.text_area.config(state='disabled')


    # MacOS弹出窗口用
    def bring_to_front(self):
        self.root.lift()
        self.root.focus_force()
        self.root.call('wm', 'attributes', '.', '-topmost', '1')
        self.root.call('wm', 'attributes', '.', '-topmost', '0')


    # 导入科目余额表子窗口
    def account_balance(self):

        option_blank = []
        option_model = ['借方列贷方列双列模式', '借贷方向列借贷金额列模式']

        top = tk.Toplevel(self.root)
        top.title('导入科目余额表')
        top.geometry('680x340+100+100')
        top.resizable(False, False)

        frame_1 = tk.Frame(top)
        frame_1.pack(side=tk.TOP, fill=tk.BOTH, pady=(5, 0))
        tk.Label(frame_1, text='File Path', width=10, anchor='w').pack(side=tk.LEFT, padx=(10, 5), pady=5)
        frame_1.entry_file = tk.Entry(frame_1, state='readonly', readonlybackground='white')
        frame_1.entry_file.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 10), pady=5)

        frame_2 = tk.Frame(top)
        frame_2.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Label(frame_2, text='选择工作表', width=10, anchor='w').pack(side=tk.LEFT, padx=(10, 5), pady=5)
        frame_2.combobox_sheet = ttk.Combobox(frame_2, values=option_blank, state='readonly', width=20)
        frame_2.combobox_sheet.pack(side=tk.LEFT, padx=(5, 20), pady=5)
        tk.Label(frame_2, text='期初期末列模式', width=10, anchor='w').pack(side=tk.LEFT, padx=(20, 5), pady=5)
        frame_2.combobox_model = ttk.Combobox(frame_2, values=option_model, state='readonly', width=20)
        frame_2.combobox_model.set(option_model[0])
        frame_2.combobox_model.pack(side=tk.LEFT, padx=(5, 10), pady=5)

        frame_3 = tk.Frame(top)
        frame_3.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Label(frame_3, text='科目编码列', width=10, anchor='w').pack(side=tk.LEFT, padx=(10, 5), pady=5)
        frame_3.combobox_num = ttk.Combobox(frame_3, values=option_blank, state='readonly', width=20)
        frame_3.combobox_num.pack(side=tk.LEFT, padx=(5, 20), pady=5)
        tk.Label(frame_3, text='科目名称列', width=10, anchor='w').pack(side=tk.LEFT, padx=(20, 5), pady=5)
        frame_3.combobox_name = ttk.Combobox(frame_3, values=option_blank, state='readonly', width=20)
        frame_3.combobox_name.pack(side=tk.LEFT, padx=(5, 10), pady=5)

        frame_4 = tk.Frame(top)
        frame_4.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Label(frame_4, text='期初借方列', width=10, anchor='w').pack(side=tk.LEFT, padx=(10, 5), pady=5)
        frame_4.combobox_begin_debit = ttk.Combobox(frame_4, values=option_blank, state='readonly', width=20)
        frame_4.combobox_begin_debit.pack(side=tk.LEFT, padx=(5, 20), pady=5)
        tk.Label(frame_4, text='期初贷方列', width=10, anchor='w').pack(side=tk.LEFT, padx=(20, 5), pady=5)
        frame_4.combobox_begin_credit = ttk.Combobox(frame_4, values=option_blank, state='readonly', width=20)
        frame_4.combobox_begin_credit.pack(side=tk.LEFT, padx=(5, 10), pady=5)

        frame_5 = tk.Frame(top)
        frame_5.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Label(frame_5, text='期初方向列', width=10, anchor='w').pack(side=tk.LEFT, padx=(10, 5), pady=5)
        frame_5.combobox_begin_dc = ttk.Combobox(frame_5, values=option_blank, state='readonly', width=20)
        frame_5.combobox_begin_dc.pack(side=tk.LEFT, padx=(5, 20), pady=5)
        tk.Label(frame_5, text='期初金额列', width=10, anchor='w').pack(side=tk.LEFT, padx=(20, 5), pady=5)
        frame_5.combobox_begin_value = ttk.Combobox(frame_5, values=option_blank, state='readonly', width=20)
        frame_5.combobox_begin_value.pack(side=tk.LEFT, padx=(5, 10), pady=5)

        frame_6 = tk.Frame(top)
        frame_6.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Label(frame_6, text='本期借方列', width=10, anchor='w').pack(side=tk.LEFT, padx=(10, 5), pady=5)
        frame_6.combobox_this_debit = ttk.Combobox(frame_6, values=option_blank, state='readonly', width=20)
        frame_6.combobox_this_debit.pack(side=tk.LEFT, padx=(5, 20), pady=5)
        tk.Label(frame_6, text='本期贷方列', width=10, anchor='w').pack(side=tk.LEFT, padx=(20, 5), pady=5)
        frame_6.combobox_this_credit = ttk.Combobox(frame_6, values=option_blank, state='readonly', width=20)
        frame_6.combobox_this_credit.pack(side=tk.LEFT, padx=(5, 10), pady=5)

        frame_7 = tk.Frame(top)
        frame_7.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Label(frame_7, text='期末借方列', width=10, anchor='w').pack(side=tk.LEFT, padx=(10, 5), pady=5)
        frame_7.combobox_end_debit = ttk.Combobox(frame_7, values=option_blank, state='readonly', width=20)
        frame_7.combobox_end_debit.pack(side=tk.LEFT, padx=(5, 20), pady=5)
        tk.Label(frame_7, text='期末贷方列', width=10, anchor='w').pack(side=tk.LEFT, padx=(20, 5), pady=5)
        frame_7.combobox_end_credit = ttk.Combobox(frame_7, values=option_blank, state='readonly', width=20)
        frame_7.combobox_end_credit.pack(side=tk.LEFT, padx=(5, 10), pady=5)

        frame_8 = tk.Frame(top)
        frame_8.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Label(frame_8, text='期末方向列', width=10, anchor='w').pack(side=tk.LEFT, padx=(10, 5), pady=5)
        frame_8.combobox_end_dc = ttk.Combobox(frame_8, values=option_blank, state='readonly', width=20)
        frame_8.combobox_end_dc.pack(side=tk.LEFT, padx=(5, 20), pady=5)
        tk.Label(frame_8, text='期末金额列', width=10, anchor='w').pack(side=tk.LEFT, padx=(20, 5), pady=5)
        frame_8.combobox_end_value = ttk.Combobox(frame_8, values=option_blank, state='readonly', width=20)
        frame_8.combobox_end_value.pack(side=tk.LEFT, padx=(5, 10), pady=5)

        frame_9 = tk.Frame(top)
        frame_9.pack(side=tk.TOP, fill=tk.BOTH, padx=5, pady=5)

        tk.Button(frame_9, text='导入文件',
                  command=lambda: self.select_balance(frame_1.entry_file,
                                                      frame_2.combobox_sheet,
                                                      frame_3.combobox_num,
                                                      frame_3.combobox_name,
                                                      frame_4.combobox_begin_debit,
                                                      frame_4.combobox_begin_credit,
                                                      frame_5.combobox_begin_dc,
                                                      frame_5.combobox_begin_value,
                                                      frame_6.combobox_this_debit,
                                                      frame_6.combobox_this_credit,
                                                      frame_7.combobox_end_debit,
                                                      frame_7.combobox_end_credit,
                                                      frame_8.combobox_end_dc,
                                                      frame_8.combobox_end_value),
                  width=15).pack(side=tk.LEFT, padx=5, pady=5)

        tk.Button(frame_9, text='导入并预览',
                  command=lambda: self.import_review_balance(frame_1.entry_file,
                                                             frame_2.combobox_sheet,
                                                             frame_2.combobox_model,
                                                             frame_3.combobox_num,
                                                             frame_3.combobox_name,
                                                             frame_4.combobox_begin_debit,
                                                             frame_4.combobox_begin_credit,
                                                             frame_5.combobox_begin_dc,
                                                             frame_5.combobox_begin_value,
                                                             frame_6.combobox_this_debit,
                                                             frame_6.combobox_this_credit,
                                                             frame_7.combobox_end_debit,
                                                             frame_7.combobox_end_credit,
                                                             frame_8.combobox_end_dc,
                                                             frame_8.combobox_end_value),
                  width=15).pack(side=tk.LEFT, padx=5, pady=5)
    
        tk.Button(frame_9, text='关闭窗口',
                  command=top.destroy,
                  width=15).pack(side=tk.LEFT, padx=5, pady=5)

        frame_2.combobox_sheet.bind('<<ComboboxSelected>>',
                                    lambda event: self.balance_change(event,
                                                                      frame_1.entry_file,
                                                                      frame_2.combobox_sheet,
                                                                      frame_3.combobox_num,
                                                                      frame_3.combobox_name,
                                                                      frame_4.combobox_begin_debit,
                                                                      frame_4.combobox_begin_credit,
                                                                      frame_5.combobox_begin_dc,
                                                                      frame_5.combobox_begin_value,
                                                                      frame_6.combobox_this_debit,
                                                                      frame_6.combobox_this_credit,
                                                                      frame_7.combobox_end_debit,
                                                                      frame_7.combobox_end_credit,
                                                                      frame_8.combobox_end_dc,
                                                                      frame_8.combobox_end_value))

        top.transient(self.root)                    # 依附主窗口
        top.grab_set()                              # 禁止操作主窗口
        self.root.wait_window(top)                  # 等待子窗口关闭


    # 更新下拉列表框
    def balance_change(self, event, entry_file, combobox_sheet, combobox_num, combobox_name,
                       combobox_begin_debit, combobox_begin_credit, combobox_begin_dc, combobox_begin_value,
                       combobox_this_debit, combobox_this_credit,
                       combobox_end_debit, combobox_end_credit, combobox_end_dc, combobox_end_value):

        path = entry_file.get()
        sheet_name = combobox_sheet.get()

        file_extension = os.path.splitext(path)[1].lower()

        if file_extension == '.xlsx':
            df = pd.read_excel(path, sheet_name=sheet_name, engine='openpyxl')
        elif file_extension == '.xls':
            df = pd.read_excel(path, sheet_name=sheet_name, engine='xlrd')

        columns = df.columns.tolist()                           # 获取工作表的列名

        if columns:

            combobox_num['values'] = columns
            combobox_num.set(columns[0])
            combobox_num.config(state='readonly')
            
            combobox_name['values'] = columns
            combobox_name.set(columns[0])
            combobox_name.config(state='readonly')

            combobox_begin_debit['values'] = columns
            combobox_begin_debit.set(columns[0])
            combobox_begin_debit.config(state='readonly')

            combobox_begin_credit['values'] = columns
            combobox_begin_credit.set(columns[0])
            combobox_begin_credit.config(state='readonly')

            combobox_begin_dc['values'] = columns
            combobox_begin_dc.set(columns[0])
            combobox_begin_dc.config(state='readonly')

            combobox_begin_value['values'] = columns
            combobox_begin_value.set(columns[0])
            combobox_begin_value.config(state='readonly')

            combobox_this_debit['values'] = columns
            combobox_this_debit.set(columns[0])
            combobox_this_debit.config(state='readonly')

            combobox_this_credit['values'] = columns
            combobox_this_credit.set(columns[0])
            combobox_this_credit.config(state='readonly')

            combobox_end_debit['values'] = columns
            combobox_end_debit.set(columns[0])
            combobox_end_debit.config(state='readonly')

            combobox_end_credit['values'] = columns
            combobox_end_credit.set(columns[0])
            combobox_end_credit.config(state='readonly')

            combobox_end_dc['values'] = columns
            combobox_end_dc.set(columns[0])
            combobox_end_dc.config(state='readonly')

            combobox_end_value['values'] = columns
            combobox_end_value.set(columns[0])
            combobox_end_value.config(state='readonly')


    # 选择科目余额表文件
    def select_balance(self, entry_file, combobox_sheet, combobox_num, combobox_name,
                       combobox_begin_debit, combobox_begin_credit, combobox_begin_dc, combobox_begin_value,
                       combobox_this_debit, combobox_this_credit,
                       combobox_end_debit, combobox_end_credit, combobox_end_dc, combobox_end_value):

        path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx'),
                                                     ('Excel Files', '*.xls')])
        
        if path:
            entry_file.config(state='normal')
            entry_file.delete(0, tk.END)
            entry_file.insert(0, path)
            entry_file.config(state='readonly')

            sheet_file = pd.ExcelFile(path)
            sheetnames = sheet_file.sheet_names 

            if sheetnames:

                combobox_sheet['values'] = sheetnames
                combobox_sheet.set(sheetnames[0])
                combobox_sheet.config(state='readonly')

                self.balance_change(None, entry_file, combobox_sheet, combobox_num, combobox_name,
                                    combobox_begin_debit, combobox_begin_credit, combobox_begin_dc, combobox_begin_value,
                                    combobox_this_debit, combobox_this_credit,
                                    combobox_end_debit, combobox_end_credit, combobox_end_dc, combobox_end_value)


    # 导入、预览科目余额表
    def import_review_balance(self, entry_file, combobox_sheet, combobox_model, combobox_num, combobox_name,
                              combobox_begin_debit, combobox_begin_credit, combobox_begin_dc, combobox_begin_value,
                              combobox_this_debit, combobox_this_credit,
                              combobox_end_debit, combobox_end_credit, combobox_end_dc, combobox_end_value):

        path = entry_file.get()
        sheet_name = combobox_sheet.get()
        model = combobox_model.get()
        num = combobox_num.get()
        name = combobox_name.get()
        begin_debit = combobox_begin_debit.get()
        begin_credit = combobox_begin_credit.get()
        begin_dc = combobox_begin_dc.get()
        begin_value = combobox_begin_value.get()
        this_debit = combobox_this_debit.get()
        this_credit = combobox_this_credit.get()
        end_debit = combobox_end_debit.get()
        end_credit = combobox_end_credit.get()
        end_dc = combobox_end_dc.get()
        end_value = combobox_end_value.get()

        file_extension = os.path.splitext(path)[1].lower()

        if file_extension == '.xlsx':
            df = pd.read_excel(path, sheet_name=sheet_name, engine='openpyxl')
        elif file_extension == '.xls':
            df = pd.read_excel(path, sheet_name=sheet_name, engine='xlrd')

        self.df_balance[['科目编码']] = df[[num]]
        # df[num] = df[num].astype(str)
        self.df_balance[['科目名称']] = df[[name]]
        self.df_balance[['本期借方']] = df[[this_debit]]
        self.df_balance[['本期贷方']] = df[[this_credit]]

        if model == '借方列贷方列双列模式':

            self.df_balance[['期初借方']] = df[[begin_debit]]
            self.df_balance[['期初贷方']] = df[[begin_credit]]
            self.df_balance[['期末借方']] = df[[end_debit]]
            self.df_balance[['期末贷方']] = df[[end_credit]]

        else:

            self.df_balance['期初借方'] = df.loc[df[begin_dc] == '借', begin_value]
            self.df_balance['期初贷方'] = df.loc[df[begin_dc] == '贷', begin_value]
            self.df_balance['期末借方'] = df.loc[df[end_dc] == '借', end_value]
            self.df_balance['期末贷方'] = df.loc[df[end_dc] == '贷', end_value]

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_path = temp_file.name

        self.df_balance.to_excel(temp_path, index=False)

        # macOS
        if os.name == 'posix':
            os.system(f'open "{temp_path}"')

        # Windows
        elif os.name == 'nt':
            os.startfile(temp_path)


    # 导入序时账子窗口
    def account_chronological(self):

        option_blank = []

        top = tk.Toplevel(self.root)
        top.title('导入序时账')
        top.geometry('680x230+100+100')
        top.resizable(False, False)

        frame_1 = tk.Frame(top)
        frame_1.pack(side=tk.TOP, fill=tk.BOTH, pady=(5, 0))
        tk.Label(frame_1, text='File Path', width=10, anchor='w').pack(side=tk.LEFT, padx=(10, 5), pady=5)
        frame_1.entry_file = tk.Entry(frame_1, state='readonly', readonlybackground='white')
        frame_1.entry_file.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 10), pady=5)

        frame_2 = tk.Frame(top)
        frame_2.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Label(frame_2, text='选择工作表', width=10, anchor='w').pack(side=tk.LEFT, padx=(10, 5), pady=5)
        frame_2.combobox_sheet = ttk.Combobox(frame_2, values=option_blank, state='readonly', width=20)
        frame_2.combobox_sheet.pack(side=tk.LEFT, padx=(5, 20), pady=5)
        tk.Label(frame_2, text='凭证日期列', width=10, anchor='w').pack(side=tk.LEFT, padx=(20, 5), pady=5)
        frame_2.combobox_date = ttk.Combobox(frame_2, values=option_blank, state='readonly', width=20)
        frame_2.combobox_date.pack(side=tk.LEFT, padx=(5, 10), pady=5)

        frame_3 = tk.Frame(top)
        frame_3.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Label(frame_3, text='凭证字号列', width=10, anchor='w').pack(side=tk.LEFT, padx=(10, 5), pady=5)
        frame_3.combobox_num = ttk.Combobox(frame_3, values=option_blank, state='readonly', width=20)
        frame_3.combobox_num.pack(side=tk.LEFT, padx=(5, 20), pady=5)
        tk.Label(frame_3, text='摘要文本列', width=10, anchor='w').pack(side=tk.LEFT, padx=(20, 5), pady=5)
        frame_3.combobox_summary = ttk.Combobox(frame_3, values=option_blank, state='readonly', width=20)
        frame_3.combobox_summary.pack(side=tk.LEFT, padx=(5, 10), pady=5)

        frame_4 = tk.Frame(top)
        frame_4.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Label(frame_4, text='科目编码列', width=10, anchor='w').pack(side=tk.LEFT, padx=(10, 5), pady=5)
        frame_4.combobox_account_num = ttk.Combobox(frame_4, values=option_blank, state='readonly', width=20)
        frame_4.combobox_account_num.pack(side=tk.LEFT, padx=(5, 20), pady=5)
        tk.Label(frame_4, text='科目名称列', width=10, anchor='w').pack(side=tk.LEFT, padx=(20, 5), pady=5)
        frame_4.combobox_account_name = ttk.Combobox(frame_4, values=option_blank, state='readonly', width=20)
        frame_4.combobox_account_name.pack(side=tk.LEFT, padx=(5, 10), pady=5)

        frame_5 = tk.Frame(top)
        frame_5.pack(side=tk.TOP, fill=tk.BOTH)
        tk.Label(frame_5, text='借方金额列', width=10, anchor='w').pack(side=tk.LEFT, padx=(10, 5), pady=5)
        frame_5.combobox_debit = ttk.Combobox(frame_5, values=option_blank, state='readonly', width=20)
        frame_5.combobox_debit.pack(side=tk.LEFT, padx=(5, 20), pady=5)
        tk.Label(frame_5, text='贷方金额列', width=10, anchor='w').pack(side=tk.LEFT, padx=(20, 5), pady=5)
        frame_5.combobox_credit = ttk.Combobox(frame_5, values=option_blank, state='readonly', width=20)
        frame_5.combobox_credit.pack(side=tk.LEFT, padx=(5, 10), pady=5)

        frame_6 = tk.Frame(top)
        frame_6.pack(side=tk.TOP, fill=tk.BOTH, padx=5, pady=5)

        tk.Button(frame_6, text='导入文件',
                  command=lambda: self.select_chronological(frame_1.entry_file,
                                                            frame_2.combobox_sheet,
                                                            frame_2.combobox_date,
                                                            frame_3.combobox_num,
                                                            frame_3.combobox_summary,
                                                            frame_4.combobox_account_num,
                                                            frame_4.combobox_account_name,
                                                            frame_5.combobox_debit,
                                                            frame_5.combobox_credit),
                  width=15).pack(side=tk.LEFT, padx=5, pady=5)

        tk.Button(frame_6, text='导入并预览',
                  command=lambda: self.import_review_chronological(frame_1.entry_file,
                                                                   frame_2.combobox_sheet,
                                                                   frame_2.combobox_date,
                                                                   frame_3.combobox_num,
                                                                   frame_3.combobox_summary,
                                                                   frame_4.combobox_account_num,
                                                                   frame_4.combobox_account_name,
                                                                   frame_5.combobox_debit,
                                                                   frame_5.combobox_credit),
                  width=15).pack(side=tk.LEFT, padx=5, pady=5)

        tk.Button(frame_6, text='关闭窗口',
                  command=top.destroy,
                  width=15).pack(side=tk.LEFT, padx=5, pady=5)

        frame_2.combobox_sheet.bind('<<ComboboxSelected>>',
                                    lambda event: self.chronological_change(event,
                                                                            frame_1.entry_file,
                                                                            frame_2.combobox_sheet,
                                                                            frame_2.combobox_date,
                                                                            frame_3.combobox_num,
                                                                            frame_3.combobox_summary,
                                                                            frame_4.combobox_account_num,
                                                                            frame_4.combobox_account_name,
                                                                            frame_5.combobox_debit,
                                                                            frame_5.combobox_credit))

        top.transient(self.root)                    # 依附主窗口
        top.grab_set()                              # 禁止操作主窗口
        self.root.wait_window(top)                  # 等待子窗口关闭


    # 更新下拉列表框
    def chronological_change(self, event, entry_file, combobox_sheet, combobox_date, combobox_num, combobox_summary,
                             combobox_account_num, combobox_account_name, combobox_debit, combobox_credit):

        path = entry_file.get()
        sheet_name = combobox_sheet.get()

        file_extension = os.path.splitext(path)[1].lower()

        if file_extension == '.xlsx':
            df = pd.read_excel(path, sheet_name=sheet_name, engine='openpyxl')
        elif file_extension == '.xls':
            df = pd.read_excel(path, sheet_name=sheet_name, engine='xlrd')

        columns = df.columns.tolist()                           # 获取工作表的列名

        if columns:

            combobox_date['values'] = columns
            combobox_date.set(columns[0])
            combobox_date.config(state='readonly')

            combobox_num['values'] = columns
            combobox_num.set(columns[0])
            combobox_num.config(state='readonly')
            
            combobox_summary['values'] = columns
            combobox_summary.set(columns[0])
            combobox_summary.config(state='readonly')

            combobox_account_num['values'] = columns
            combobox_account_num.set(columns[0])
            combobox_account_num.config(state='readonly')

            combobox_account_name['values'] = columns
            combobox_account_name.set(columns[0])
            combobox_account_name.config(state='readonly')

            combobox_debit['values'] = columns
            combobox_debit.set(columns[0])
            combobox_debit.config(state='readonly')

            combobox_credit['values'] = columns
            combobox_credit.set(columns[0])
            combobox_credit.config(state='readonly')


    # 选择序时账文件
    def select_chronological(self, entry_file, combobox_sheet, combobox_date, combobox_num, combobox_summary,
                             combobox_account_num, combobox_account_name, combobox_debit, combobox_credit):

        path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx'),
                                                     ('Excel Files', '*.xls')])
        
        if path:
            entry_file.config(state='normal')
            entry_file.delete(0, tk.END)
            entry_file.insert(0, path)
            entry_file.config(state='readonly')

            sheet_file = pd.ExcelFile(path)
            sheetnames = sheet_file.sheet_names 

            if sheetnames:

                combobox_sheet['values'] = sheetnames
                combobox_sheet.set(sheetnames[0])
                combobox_sheet.config(state='readonly')

                self.chronological_change(None, entry_file, combobox_sheet, combobox_date, combobox_num, combobox_summary,
                                          combobox_account_num, combobox_account_name, combobox_debit, combobox_credit)


    # 导入、预览序时账
    def import_review_chronological(self, entry_file, combobox_sheet, combobox_date, combobox_num, combobox_summary,
                                    combobox_account_num, combobox_account_name, combobox_debit, combobox_credit):

        path = entry_file.get()
        sheet_name = combobox_sheet.get()
        date = combobox_date.get()
        number = combobox_num.get()
        summary = combobox_summary.get()
        account_num = combobox_account_num.get()
        account_name = combobox_account_name.get()
        debit = combobox_debit.get()
        credit = combobox_credit.get()

        file_extension = os.path.splitext(path)[1].lower()

        if file_extension == '.xlsx':
            df = pd.read_excel(path, sheet_name=sheet_name, engine='openpyxl')
        elif file_extension == '.xls':
            df = pd.read_excel(path, sheet_name=sheet_name, engine='xlrd')

        # 对借方、贷方发生额进行数据清洗
        df[debit] = df[debit].apply(self.convert_to_numeric)
        df[credit] = df[credit].apply(self.convert_to_numeric)

        # 日期列列填充空行（重复上一行），转换制单日期为datetime类型并提取月份
        df[date] = df[date].ffill()
        df[date] = pd.to_datetime(df[date], errors='coerce')
        df['月份'] = df[date].dt.month

        # 凭证号列填充空行（重复上一行）
        df[number] = df[number].ffill()

        # 遍历每个凭证号和月份，找出对应的科目
        voucher_groups = df.groupby(['月份', number])
        for (month, voucher_no), group in voucher_groups:
            primary_subjects = set()
            for subject in group[account_num]:
                primary_subjects.add(str(subject))
            primary_subjects_str = ', '.join(primary_subjects)
            df.loc[(df['月份'] == month) & (df[number] == voucher_no), '涉及科目'] = primary_subjects_str

        self.df_chronological['涉及科目'] = df['涉及科目']
        self.df_chronological['日期'] = df[date]
        self.df_chronological['凭证字号'] = df[number]
        self.df_chronological['摘要'] = df[summary]
        self.df_chronological['科目编码'] = df[account_num]
        self.df_chronological['科目名称'] = df[account_name]
        self.df_chronological['借方金额'] = df[debit]
        self.df_chronological['贷方金额'] = df[credit]

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_path = temp_file.name

        self.df_chronological.to_excel(temp_path, index=False)

        # macOS
        if os.name == 'posix':
            os.system(f'open "{temp_path}"')

        # Windows
        elif os.name == 'nt':
            os.startfile(temp_path)


    # 数据清洗
    def convert_to_numeric(self, value):

        try:
            return pd.to_numeric(value.replace(',', ''))
        
        except AttributeError:
            return 0 if pd.isna(value) else value


    # 导出按钮
    def export_file(self, text_area):

        fill_text = ''

        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.file_path_result = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel Files', '*.xlsx')])

        if self.file_path_result:

            wb = openpyxl.Workbook()
            ws_find = wb.active
            ws_find.title = '查找'

            cell_A1 = ws_find.cell(row=1, column=1, value='科目代码')
            cell_A1.border = thin_border

            cell_B1 = ws_find.cell(row=1, column=2, value='')
            cell_B1.number_format = '@'
            cell_B1.border = thin_border

            ws_find.cell(row=3, column=1, value='科目余额表')               # A3

            cell_A4 = ws_find.cell(row=4, column=1, value='科目编码')       # A4
            cell_A4.number_format = '@'

            ws_find.cell(row=4, column=2, value='科目名称')                 # B4
            ws_find.cell(row=4, column=3, value='期初借方')                 # C4
            ws_find.cell(row=4, column=4, value='期初贷方')                 # D4
            ws_find.cell(row=4, column=5, value='本期借方')                 # E4
            ws_find.cell(row=4, column=6, value='本期贷方')                 # F4
            ws_find.cell(row=4, column=7, value='期末借方')                 # G4
            ws_find.cell(row=4, column=8, value='期末贷方')                 # H4

            for row in ws_find.iter_rows(min_row=4, max_row=5, min_col=1, max_col=8):
                for cell in row:
                    cell.border = thin_border

            ws_find.cell(row=7, column=1, value='序时账')                  # A7

            ws_find.cell(row=7, column=1, value='涉及科目')                 # A7
            ws_find.cell(row=7, column=2, value='日期')                    # B7
            ws_find.cell(row=7, column=3, value='凭证字号')                 # C7
            ws_find.cell(row=7, column=4, value='科目编码')                 # D7
            ws_find.cell(row=7, column=5, value='科目名称')                 # E7
            ws_find.cell(row=7, column=6, value='摘要')                    # F7
            ws_find.cell(row=7, column=7, value='借方金额')                 # G7
            ws_find.cell(row=7, column=8, value='贷方金额')                 # H7

            for row in ws_find.iter_rows(min_row=7, max_row=10007, min_col=1, max_col=8):
                for cell in row:
                    cell.border = thin_border

            ws_find.auto_filter.ref = "A7:H7"

            # 设置日期和数值格式，遍历所有列
            for col in ws_find.iter_cols(min_row=8, max_row=10007, min_col=2, max_col=2):
                for cell in col:
                    cell.number_format = 'yyyy-mm-dd'

            for col in ws_find.iter_cols(min_row=8, max_row=10007, min_col=7, max_col=8):
                for cell in col:
                    cell.number_format = '#,##0.00'             # 数字格式设置为千分位和保留两位小数

            # 调整列宽（按实际列顺序）
            find_widths = {
                            'A': 15,
                            'B': 15,
                            'C': 15,
                            'D': 25,
                            'E': 50,
                            'F': 50,
                            'G': 15,
                            'H': 15
            }

            # 设置列宽
            for i, (col_name, width) in enumerate(find_widths.items(), start=1):
                col_letter = get_column_letter(i)
                ws_find.column_dimensions[col_letter].width = width


            ws_balance = wb.create_sheet(title='科目余额表')

            for j, col in enumerate(self.df_balance.columns):
                ws_balance.cell(row=1, column=1+j, value=col)

            for i, row in enumerate(self.df_balance.values):
                for j, value in enumerate(row):
                    ws_balance.cell(row=2+i, column=1+j, value=value)

            for row in ws_balance.iter_rows(min_row=1, max_row=ws_balance.max_row, min_col=1, max_col=ws_balance.max_column):
                for cell in row:
                    cell.border = thin_border

            # 添加筛选功能，筛选行设置在第一行
            ws_balance.auto_filter.ref = ws_balance.dimensions

            # 设置日期和数值格式，遍历所有列
            for col in ws_balance.iter_cols(min_row=2, max_row=ws_balance.max_row, min_col=2, max_col=ws_balance.max_column):
                for cell in col:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'             # 数字格式设置为千分位和保留两位小数

            # 调整列宽（按实际列顺序）
            balance_widths = {
                '科目编号': 15,
                '科目名称': 50,
                '期初借方': 15,
                '期初贷方': 15,
                '本期借方': 15,
                '本期贷方': 15,
                '期末借方': 15,
                '期末贷方': 15
            }

            # 设置列宽
            for i, (col_name, width) in enumerate(balance_widths.items(), start=1):
                col_letter = get_column_letter(i)
                ws_balance.column_dimensions[col_letter].width = width


            ws_chronological = wb.create_sheet(title='序时账')

            for j, col in enumerate(self.df_chronological.columns):
                ws_chronological.cell(row=1, column=1+j, value=col)

            for i, row in enumerate(self.df_chronological.values):
                for j, value in enumerate(row):
                    ws_chronological.cell(row=2+i, column=1+j, value=value)

            for row in ws_chronological.iter_rows(min_row=1, max_row=ws_chronological.max_row, min_col=1, max_col=ws_chronological.max_column):
                for cell in row:
                    cell.border = thin_border

            # 添加筛选功能，筛选行设置在第一行
            ws_chronological.auto_filter.ref = ws_chronological.dimensions

            # 设置日期和数值格式，遍历所有列
            for col in ws_chronological.iter_cols(min_row=2, max_row=ws_chronological.max_row, min_col=2, max_col=2):
                for cell in col:
                    cell.number_format = 'yyyy-mm-dd'

            for col in ws_chronological.iter_cols(min_row=2, max_row=ws_chronological.max_row, min_col=7, max_col=ws_chronological.max_column):
                for cell in col:
                    cell.number_format = '#,##0.00'             # 数字格式设置为千分位和保留两位小数

            # 调整列宽（按实际列顺序）
            chronological_widths = {
                                    '涉及科目': 10,
                                    '日期': 15,
                                    '凭证字号': 15,
                                    '科目编码': 30,
                                    '科目名称': 50,
                                    '摘要': 50,
                                    '借方金额': 15,
                                    '贷方金额': 15
            }

            # 设置列宽
            for i, (col_name, width) in enumerate(chronological_widths.items(), start=1):
                col_letter = get_column_letter(i)
                ws_chronological.column_dimensions[col_letter].width = width

            wb.save(self.file_path_result)

            subprocess.run(['open', self.file_path_result])

            fill_text += 'A5单元格公式：=XLOOKUP(B1,科目余额表!A:A,科目余额表!A:H)\n'
            fill_text += 'A8单元格公式：=FILTER(序时账!A:H,ISNUMBER(SEARCH(B1,序时账!A:A)))\n'
            fill_text += f'Path: {self.file_path_result}\nExport successful!'

            text_area.config(state='normal')                                        # 临时启用
            text_area.insert(tk.INSERT, fill_text)
            text_area.see(tk.END)                                                   # 滚动到底部
            text_area.config(state='disabled')                                      # 重新禁用


app = App()
app.root.mainloop()